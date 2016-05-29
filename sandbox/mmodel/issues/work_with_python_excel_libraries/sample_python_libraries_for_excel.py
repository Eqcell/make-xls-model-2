#!/usr/bin/env python3
"""
Write formula: 
+ xlwt OK
- openpyxl - converter fails in Excel 2003 in opening this file - similar to https://groups.google.com/forum/#!topic/openpyxl-users/f2LNafTUKwY
- xlsxwriter

Read formula:
+ openpyxl 
- xlrd
- xlsxwriter

... but writing this in 'xlwings' is easiest.
"""

import pandas as pd
df = pd.DataFrame([500.5, "=SUM(A1, -200)"])
df.to_excel("df.xlsx")


###############################################################################
# openpyxl 
from openpyxl import Workbook, load_workbook
wb = Workbook()
ws = wb.active
# add a simple formula
ws["A1"] = "=SUM(1, 1)"
ws["A2"] = 500.5
wb.save("f_openpyxl.xlsx")

zb = load_workbook(filename="formula_openpyxl.xlsx")
ws = zb['Sheet'] 
for row in ws.rows:
    for cell in row:
        print(cell.value)

###############################################################################
# xlwt/xlrd
import xlwt
from datetime import datetime
wb = xlwt.Workbook()
ws = wb.add_sheet('sheet_1')

ws.write(0, 0, 1234.56)
ws.write(1, 0, datetime.now())
ws.write(2, 0, 1)
ws.write(2, 1, 1)
ws.write(2, 2, xlwt.Formula("A3+B3"))
ws.write(3, 2, xlwt.Formula("SUM(100, 11)"))
wb.save("f_xlwt.xls")

###############################################################################        
# xlsxwriter
import xlsxwriter

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook("f_xlsxwriter.xlsx")
worksheet = workbook.add_worksheet()

# Some data we want to write to the worksheet.
expenses = (
    ['Rent', 1000],
    ['Gas',   100],
    ['Food',  300],
    ['Gym',    50],
)

# Start from the first cell. Rows and columns are zero indexed.
row = 0
col = 0

# Iterate over the data and write it out row by row.
for item, cost in (expenses):
    worksheet.write(row, col,     item)
    worksheet.write(row, col + 1, cost)
    row += 1

# Write a total using a formula.
worksheet.write(row, 0, 'Total')
# line below will not write formula
# worksheet.write(row, 1, '=SUM(B1:B4)')
worksheet.write_formula('B6', '=SUM(B1:B4)', value = 1000 + 100 + 300 + 50)

workbook.close()

###############################################################################        
#xlwings

#from xlwings import Workbook, Range
#wb = Workbook(REF_FILE)
#z = Range('result', 'A1').table.value
#print(z)
