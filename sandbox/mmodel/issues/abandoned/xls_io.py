import sys
import os
import pandas as pd
import numpy as np
from pprint import pprint
from datetime import datetime

import xlsxwriter
from openpyxl import Workbook
from xlwt import Workbook as xlwtWorkbook, Formula


    
def write_output_to_xls(ar, view_spec):

    # NEXT: change to view_spec parameters, unpack them
    sheet_name = "model"
    xlsx_path = "D:/make-xls-model-master/spec.xls"
    write_array_to_xlsx_using_xlwings(ar, xlsx_path, sheet_name)  
    
    # if method is None:
        # write_array_to_xlsx_using_xlwings(ar, xlsx_path, sheet_name)    
    # if method == 'xlwt' or method is None:        
        # write_array_to_xlsx_using_xlwt(ar, xlsx_path, sheet_name)
    # elif method == 'openpyxl':
        # xlsx_path = replace_extension(xlsx_path)
        # write_array_to_xlsx_using_openpyxl(ar, xlsx_path, sheet_name)
    # elif method == 'xlsxwriter':    
        # xlsx_path = replace_extension(xlsx_path)
        # write_array_to_xlsx_using_xlsxwriter(ar, xlsx_path, sheet_name)
    # else:
        # raise ValueError("Method for writing Excel file not defined: " + method)
        
############# iterators ################## 

def to_float(a):
    a = str(a)
    try:
        z = float(a)
        if round(z) == z:
           return int(z)
        else:
           return z
    except ValueError:
        return a
        
def iterate_over_array(ar):
    for i, row in enumerate(ar):       
         for j, val in enumerate(row):
                yield i, j, to_float(val)
                
def finished_writing(xlsx_path):
    print ("Finished_writing:", xlsx_path)

    
    
###############   xlwings    ###############  
from xlwings import Workbook, Range, Sheet
   
def write_array_to_xlsx_using_xlwings(ar, file, sheet):    

    wb = Workbook(file)
    Sheet(sheet).activate()    
    for i, j, val in iterate_over_array(ar):
        Range(sheet, (i, j)).value = val  
    wb.save()
    
    # WARNING: 
    #    Range(sheet, 'A1').value = ar
    #    works with error  
    
  
    
    
    
###############   xlwt    ###############  

def write_array_to_xlsx_using_xlwt(ar, xlsx_path, sheet_name):
    book = xlwtWorkbook()
    ws = book.add_sheet(sheet_name)

    for i, j, val in iterate_over_array(ar):
       if str(val).startswith("="):            
            ws.write(i,j,Formula(val[1:]))
       else:                    
            ws.write(i,j,val)
            
    # http://www.simplistix.co.uk/presentations/python-excel.pdf
    # sheet1.write(0,1,'B1')
    # row1 = sheet1.row(1)
    # row1.write(0,'A2')
    # row1.write(1,'B2')
    # sheet1.col(0).width = 10000
    # sheet2 = book.get_sheet(1)
    # sheet2.row(0).write(0,'Sheet 2 A1')
    # sheet2.row(0).write(1,'Sheet 2 B1')
    # sheet2.flush_row_data()
    # sheet2.write(1,0,'Sheet 2 A3')
    # sheet2.col(0).width = 5000
    # sheet2.col(0).hidden = True
    
    book.save(xlsx_path)
    finished_writing(xlsx_path)
    
                
############### openpyxl ###############  
                
def write_array_to_xlsx_using_openpyxl(ar, xlsx_path, sheet_name):
    wb = Workbook()
    ws = wb.active # wb.create_sheet(0, sheet_name)
    
    for i, j, val in iterate_over_array(ar):
                if str(val).startswith("="):
                     ws.cell(row = i, column = j).value = val
                else:                    
                     ws.cell(row = i, column = j).value = val
                     
    wb.save(xlsx_path)
    finished_writing(xlsx_path)

    
############### xlsxwriter ###############  
    
def write_array_to_xlsx_using_xlsxwriter(ar, xlsx_path, sheet_name):
    
    workbook = xlsxwriter.Workbook(xlsx_path)
    ws = workbook.add_worksheet(sheet_name)
    cell_format = get_cell_format(workbook)  
    
    for i, j, val in iterate_over_array(ar):                
                if str(val).startswith("="):
                     ws.write_formula(i, j, val) #, cell_format, 0)
                else:                    
                      ws.write(i, j, val)
    workbook.close()
    finished_writing(xlsx_path)
    

def get_cell_format(workbook):
    """
    Cell format dictionary. 
    """   
    format_ = workbook.add_format()
    format_.set_font_name('Arial')
    format_.set_font_size(8)
    return format_
    
def get_date_format(workbook):
    """
    Date format dictionary. 
    """ 
    format_ = get_cell_format(workbook)
    format_.set_num_format('dd.mm.yyyy')
    return format_
    
############### ******** ###############  
