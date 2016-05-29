import pandas as pd
import numpy as np
import os
from pprint import pprint
from xlwings import Workbook, Range, Sheet
from openpyxl import load_workbook

from iterate_in_array import fill_array_with_excel_formulas
from iterate_in_array import fill_array_with_excel_formulas_based_on_is_forecast   

from import_specification import get_all_input_variables, get_array_and_support_variables, get_dataset_df, validate_variable_names



###########################################################################
## Export to Excel workbook (using xlwings/pywin32 or openpyxl)
###########################################################################

def save_xl_using_xlwings(file):  
    wb = Workbook(file)
    wb.save()

def write_array_to_xl_using_xlwings(ar, file, sheet):
    # Note: if file is opened In Excel, it must be first saved before writing 
    #       new output to it, but it may be left open in Excel application. 
    wb = Workbook(file)
    Sheet(sheet).activate()

    def nan_to_empty_str(x):
        return '' if type(x) == float and np.isnan(x) else x

    Range(sheet, 'A1').value = [[nan_to_empty_str(x) for x in row] for row in ar]
    wb.save()

#--------------------------------------------------------------------------
# Not tested below: alternative Excel writer

def change_extension(file):
    """
    >>> change_extension("spec.xls")
    'spec.xlsx'
    """
    return os.path.splitext(file) + ".xlsx"

def iterate_over_array(ar):
    for i, row in enumerate(ar):       
         for j, val in enumerate(row):
                yield i, j, val
                
def write_array_to_xlsx_using_openpyxl(ar, file, sheet):  
    wb = load_workbook(file)
    ws = wb.get_sheet_by_name(sheet)
    for i, j, val in iterate_over_array(ar):
        ws.cell(row = i, column = j).value = val
    new_filename = change_extension(file)
    wb.save(new_filename) 
    
#--------------------------------------------------------------------------
    
###########################################################################
## Dataframe manipulation
###########################################################################

def make_empty_df(index_, columns_):
    df = pd.DataFrame(index=index_, columns=columns_)
    return  df 

def subset_df(df, var_list):
    try:
        return df[var_list]
    except KeyError:        
        pprint ([x for x in var_list if x not in df.columns.values])
        raise KeyError("*var_list* contains variables outside *df* column names." +  
                       "\nCannot perform subsetting like df[var_list]")
    except:
        print ("Error handling dataframe:", df)
        raise ValueError
      
def make_df_before_equations(data_df, controls_df, equations_dict, var_group):
    """
    Return a dataframe containing data, controls and a placeholder for new 
    variables derived in equations.
    """    
    IS_FORECAST_LABEL = 'is_forecast'
    
    # assign 'is_forecast' to dataframes
    data_df[IS_FORECAST_LABEL] = 0 
    controls_df[IS_FORECAST_LABEL] = 1
     
    # concat data and control *df*
    df = data_df.combine_first(controls_df)
    
    
    # *df2* is a placeholder for equation-derived variables 
    df3 = make_empty_df(data_df.index, var_group['eq'])
    # add *df3* to *df*. 
    df = pd.merge(df, df3, left_index = True, right_index = True, how = 'left')
    
    # reorganise rows
    var_list = var_group['data'] +  var_group['eq'] + var_group['control'] + [IS_FORECAST_LABEL]
    return subset_df(df, var_list)


###########################################################################
## Array manipulations
###########################################################################

from iterate_in_array import get_variable_rows_as_dict

def make_array_before_equations(df):
    """
    Convert dataframe to array, decorate with extra top row an extra left-side columns.
    Returns array and pivot column number. Pivot column contains variable labels.
    """
    ar = df.as_matrix().transpose().astype(object)
    
    # add variable labels as a first column in *ar*
    labels = df.columns.tolist()
    ar = np.insert(ar, 0, labels, axis = 1)
    pivot_col = 0

    # add years as first row in *ar*
    years = [""] + df.index.astype(str).tolist()
    ar = np.insert(ar, 0, years, axis = 0)
    
    return ar, pivot_col

###### After equations

def insert_empty_row_before_variable(ar, var_name, pivot_col, start_cell_text = ""):
    variables_dict = get_variable_rows_as_dict(ar, pivot_col)
    row_position = variables_dict[var_name] 
    ar = np.insert(ar, row_position, "", axis = 0) 
    ar[row_position, 0] = start_cell_text 
    return ar

def insert_column(ar, pivot_col, datagen_func):
    column_values = [datagen_func(x) for x in ar[:, pivot_col]]
    ar = np.insert(ar, 0, column_values, axis = 1)
    return ar, pivot_col + 1   

def append_row_to_array(ar):
    row = [["" for x in ar[0,:]]]    
    return np.append(ar, row, axis = 0)
    
def add_equations_to_array (ar, pivot_col, eq_list):    
    for eq in eq_list:
        ar = append_row_to_array(ar)
        ar[-1, pivot_col] = eq
    return ar

###########################################################################
## Split *dataset*  sheet
###########################################################################

def dataset_to_basic_sheets(abs_filepath):
    dataset = get_dataset_df(abs_filepath)

    def remove_useless_columns(df):
        columns_to_remove = [x for x in df.columns if 'Unnamed:' in str(x)]
        return df.drop(columns_to_remove, 1)

    def remove_rows_with_no_type(df):
        return df[df.type.notnull()]

    dataset = remove_useless_columns(dataset)
    dataset = remove_rows_with_no_type(dataset)

    names = dataset[dataset.type.map(lambda x: x in ['data', 'param'])][['var', 'name']]
    validate_variable_names(names['var'].tolist())

    equations = dataset[dataset.type == 'eq'][['name']]

    def get_data_years(df):
        years = df[dataset.type == 'is_forecast'].iloc[0]
        return [year for year, is_forecast in years.iteritems() if is_forecast == 0]

    data = dataset[dataset.type == 'data'][['var'] + get_data_years(dataset)]
    controls = dataset[dataset.type == 'param'].drop(['type', 'name'], 1)

    def to_array_with_years_and_values(df):
        result = np.concatenate(([np.array(df.columns)], df.as_matrix().astype(object)))
        result[0][0] = ''
        return result

    write_array_to_xl_using_xlwings(to_array_with_years_and_values(data), abs_filepath, 'data')
    write_array_to_xl_using_xlwings(to_array_with_years_and_values(controls), abs_filepath, 'controls')
    write_array_to_xl_using_xlwings(names.as_matrix(), abs_filepath, 'names')
    write_array_to_xl_using_xlwings(equations.as_matrix(), abs_filepath, 'equations')


###########################################################################
## Main fucntional entry points (all of this section to be replaced with higher level classes?)
###########################################################################

# todo: function below will be replaced by higher level classes?
def get_resulting_workbook_array_for_make(abs_filepath, slim = True):

    # Get model specification
    data_df, controls_df, equations_dict, var_group, var_desc_dict, eq_list = get_all_input_variables(abs_filepath) 
     
    # Get array before formulas
    df = make_df_before_equations(data_df, controls_df, equations_dict, var_group)
    ar, pivot_col = make_array_before_equations(df) 
    
    if not slim:
        # Decorate with extra columns ---------------------------------------------
        def null(x):    
           return ""
           
        def get_var_desc(varname):
           if varname in var_desc_dict.keys():
               return var_desc_dict[varname]
           else:
               return ""       
           
        ar, pivot_col = insert_column(ar, pivot_col, get_var_desc)
        ar, pivot_col = insert_column(ar, pivot_col, null)              
       
        # Decorate with extra empty rows 
        def insert_row(t, gen):
            # t is (varname, start_cell_text)
            return insert_empty_row_before_variable(ar, t[0], 
                                                    pivot_col, next(gen) + t[1])        
        def yield_chapter_numbers():
            for i in [1,2,3,4]:
                 yield str(i)  
                                                   
        gen = yield_chapter_numbers()
        
        dec_dict = { "data": (var_group['data'][0],    ". ИСХОДНЫЕ ДАННЫЕ И ПРОГНОЗ"),
                     "ctrl": (var_group['control'][0], ". УПРАВЛЯЮЩИЕ ПАРАМЕТРЫ")}                      
        if var_group['eq']:
             dec_dict['eq'] = (var_group['eq'][0],      ". ПЕРЕМЕННЫЕ ИЗ УРАВНЕНИЙ")
                        
        ar = insert_row(dec_dict['data'], gen)
        if var_group['eq']:        
            ar = insert_row(dec_dict['eq'], gen)
        ar = insert_row(dec_dict['ctrl'], gen)
                        
        # -------------------------------------------------------------------------    
       
    ar = fill_array_with_excel_formulas(ar, equations_dict, pivot_col)
    
    if not slim:
        ar = append_row_to_array(ar)
        ar[-1,0] = next(gen) + ". УРАВНЕНИЯ"
        ar = add_equations_to_array (ar, pivot_col, eq_list)
        
    return ar

# pivot_col = 2 is standard output of --make --fancy
# todo: function below will be replaced by higher level classes?
def update_xl_model(abs_filepath, sheet, pivot_col = 2): 
    save_xl_using_xlwings(abs_filepath) 
    ar, equations_dict = get_array_and_support_variables(abs_filepath, sheet, pivot_col)         
    ar = fill_array_with_excel_formulas_based_on_is_forecast(ar, equations_dict, pivot_col)    
    print("\nResulting Excel sheet as array:")     
    print(ar)  
    write_array_to_xl_using_xlwings(ar, abs_filepath, sheet)

# todo: is this a duplicate that is never used? delete/comment out?
def derive_sheets_from_dataset(abs_filepath):
    dataset_to_basic_sheets(abs_filepath)

# todo: function below will be replaced by higher level classes?
def make_xl_model(abs_filepath, sheet, slim):
    ar = get_resulting_workbook_array_for_make(abs_filepath, slim)
    print("\nResulting Excel sheet as array:")     
    print(ar)
    write_array_to_xl_using_xlwings(ar, abs_filepath, sheet)

###########################################################################
## Higher level classes - new entry points, used in model.py
###########################################################################

class ExcelFileWorker:

    def __init__(self, excel_file):
        self.file = excel_file

        if not os.path.isfile(self.file):
            raise Exception("ERROR: file {} doesn't exist".format(self.file))

    def _save_array(self, ar, sheet):
        write_array_to_xl_using_xlwings(ar, self.file, sheet)


class _Model(ExcelFileWorker):
    
    def __init__(self, excel_file, model_sheet):
        ExcelFileWorker.__init__(self, excel_file)
        self.model_sheet = model_sheet
        self.model_array = None
        
    def save(self):
        if self.model_array is not None and self.model_sheet is not None:
            self._save_array(self.model_array, self.model_sheet)
        else:
            raise Exception("In class _Model: cannot save model because 'self.model_array' and 'self.model_sheet' are not defined")

    def print_model_sheet(self):
        if self.model_array is not None:
            print("\nResulting Excel sheet as array:")
            print(self.model_array)
        else:
            raise Exception("In class _Model: cannot print model because 'self.model_array' is not defined")


class DatasetSplitter(ExcelFileWorker):

    def __init__(self, excel_file):
        ExcelFileWorker.__init__(self, excel_file)
        self.names, self.equations, self.data, self.controls = None, None, None, None

    def derive_from_dataset(self):
        dataset = get_dataset_df(self.file)

        def remove_useless_columns(df):
            columns_to_remove = [x for x in df.columns if 'Unnamed:' in str(x)]
            return df.drop(columns_to_remove, 1)

        def remove_rows_with_no_type(df):
            return df[df.type.notnull()]

        dataset = remove_useless_columns(dataset)
        dataset = remove_rows_with_no_type(dataset)

        self.names = dataset[dataset.type.map(lambda x: x in ['data', 'param'])][['var', 'name']]
        validate_variable_names(self.names['var'].tolist())

        self.equations = dataset[dataset.type == 'eq'][['name']]

        def get_data_years(df):
            years = df[dataset.type == 'is_forecast'].iloc[0]
            return [year for year, is_forecast in years.iteritems() if is_forecast == 0]

        self.data = dataset[dataset.type == 'data'][['var'] + get_data_years(dataset)]
        self.controls = dataset[dataset.type == 'param'].drop(['type', 'name'], 1)

    def save(self):
        def to_array_with_years_and_values(df):
            result = np.concatenate(([np.array(df.columns)], df.as_matrix().astype(object)))
            result[0][0] = ''
            return result

        write_array_to_xl_using_xlwings(to_array_with_years_and_values(self.data), self.file, 'data')
        write_array_to_xl_using_xlwings(to_array_with_years_and_values(self.controls), self.file, 'controls')
        write_array_to_xl_using_xlwings(self.names.as_matrix(), self.file, 'names')
        write_array_to_xl_using_xlwings(self.equations.as_matrix(), self.file, 'equations')


class ModelCreator(_Model):

    def __init__(self, excel_file, model_sheet):
        _Model.__init__(self, excel_file, model_sheet)

        self.model_df = None
        self.pivot_col = None
        
        # question: maybe this should not be global to class instance, but rather local 
        #           for _add_descriptions_for_all_except_equations(self)? or there is reason why it is better here?
        def yield_chapter_numbers():
            for i in [1,2,3,4]:
                 yield str(i)
        self.gen = yield_chapter_numbers()

    def build_slim(self):
        self._load_main_sheets_to_array()
        self._fill_array_with_excel_formulas()

    def build_fancy(self):
        self._load_main_sheets_to_array()
        self._add_descriptions_for_all_except_equations()
        self._fill_array_with_excel_formulas()
        self._add_description_for_equations()

    def _load_main_sheets_to_array(self):
        self._read_main_sheets()
        self._make_dataframe()
        self._make_array_with_data_and_controls()

    def _read_main_sheets(self):
        self.data_df, self.controls_df, self.equations_dict, self.var_group, self.var_desc_dict, self.eq_list =\
            get_all_input_variables(self.file)

    def _make_dataframe(self):
        self.model_df = make_df_before_equations(self.data_df, self.controls_df, self.equations_dict, self.var_group)

    def _make_array_with_data_and_controls(self):
        self.model_array, self.pivot_col = make_array_before_equations(self.model_df)

    def _add_descriptions_for_all_except_equations(self):
        def null(x):
           return ""

        def get_var_desc(varname):
           if varname in self.var_desc_dict.keys():
               return self.var_desc_dict[varname]
           else:
               return ""

        self.model_array, self.pivot_col = insert_column(self.model_array, self.pivot_col, get_var_desc)
        self.model_array, self.pivot_col = insert_column(self.model_array, self.pivot_col, null)

        # Decorate with extra empty rows
        def insert_row(t, gen):
            # t is (varname, start_cell_text)
            return insert_empty_row_before_variable(self.model_array, t[0],
                                                    self.pivot_col, next(gen) + t[1])

        dec_dict = {
            "data": (self.var_group['data'][0],    ". ИСХОДНЫЕ ДАННЫЕ И ПРОГНОЗ"),
            "ctrl": (self.var_group['control'][0], ". УПРАВЛЯЮЩИЕ ПАРАМЕТРЫ")
        }
        if self.var_group['eq']:
             dec_dict['eq'] = (self.var_group['eq'][0],      ". ПЕРЕМЕННЫЕ ИЗ УРАВНЕНИЙ")

        self.model_array = insert_row(dec_dict['data'], self.gen)
        if self.var_group['eq']:
            self.model_array = insert_row(dec_dict['eq'], self.gen)
        self.model_array = insert_row(dec_dict['ctrl'], self.gen)

    def _add_description_for_equations(self):
        self.model_array = append_row_to_array(self.model_array)
        self.model_array[-1,0] = next(self.gen) + ". УРАВНЕНИЯ"
        self.model_array = add_equations_to_array(self.model_array, self.pivot_col, self.eq_list)

    def _fill_array_with_excel_formulas(self):
        self.model_array = fill_array_with_excel_formulas(self.model_array, self.equations_dict, self.pivot_col)


class ModelUpdater(_Model):

    def __init__(self, excel_file, model_sheet):
        _Model.__init__(self, excel_file, model_sheet)
        self.pivot_col = 2

    def update_model(self):
        save_xl_using_xlwings(self.file)
        self.model_array, equations_dict = get_array_and_support_variables(self.file, self.model_sheet, self.pivot_col)
        self.model_array = fill_array_with_excel_formulas_based_on_is_forecast(self.model_array, equations_dict, self.pivot_col)


# Comment: I also thought of an extra parent class ```DefaultSheetNames``` or ```DefaultSheetConfiguration``` 
#          holding sheet names like 'model', 'dataset', etc and some constants. Given more though I noticed 
#          it does not really fit to high level classes, but rather lower level classes. 
#          For example default names for sheets are hidden in import_specification.py
#          so this does not apply to highlevel classes, just something to keep in mind if we go further to creating more classes, not todo immediately.
